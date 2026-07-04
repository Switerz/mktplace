"""
Testes de apps/api/etl/load_shopee_products.py — endurecimento de
_clean_numeric/_parse_brl_float/ShopeeNumericParseError (2026-07-04).

Contexto: esta é uma implementação LOCAL, independente do parser
canônico em pipelines/connectors/shopee/_numeric.py::parse_brl_float
(mesmo contrato, testado separadamente em
pipelines/tests/test_shopee_numeric.py) — ver docstring de
ShopeeNumericParseError em load_shopee_products.py para a justificativa
completa de por que não é importada (apps/api é empacotado e implantado
de forma independente; `import pipelines` falha no .venv real de
apps/api quando executado com cwd=apps/api).

Nenhum teste aqui abre conexão com banco nem executa main().

IMPORTANTE: este arquivo importa o MÓDULO (`from etl import
load_shopee_products as mod`) e acessa `mod.ShopeeNumericParseError`/
`mod._clean_numeric`/`mod._parse_brl_float` por atributo, em vez de
`from etl.load_shopee_products import ShopeeNumericParseError, ...` no
topo do arquivo. Motivo: test_load_shopee_products_local_pg_guard.py
chama `importlib.reload()` no módulo — um `from X import Y` feito antes
do reload fica com uma referência à classe/função ANTIGA, que não é mais
`is`-idêntica à nova definida pelo reload, e `pytest.raises(OldClasse)`
deixa de casar com uma instância da NovaClasse quando a suíte inteira
roda em conjunto (falha só reproduzível com a suíte completa, não com
este arquivo isolado — diagnosticado nesta sessão).
"""
from __future__ import annotations

import inspect
import math
import re
import traceback

import openpyxl
import pandas as pd
import pytest

from etl import load_shopee_products as mod


def _series(values, files=None, rows=None):
    n = len(values)
    files = files or [f"arquivo_{i}.xlsx" for i in range(n)]
    rows = rows or [i + 2 for i in range(n)]
    return (
        pd.Series(values),
        pd.Series(files),
        pd.Series(rows),
    )


# --- formato real confirmado nos exports (ponto decimal, sem milhar) ---
# Amostras numéricas anonimizadas (não são PII) já auditadas em
# pipelines/tests/test_shopee_numeric.py — mesmos valores, fonte
# equivalente ("Subtotal do produto").

@pytest.mark.parametrize("raw, expected", [
    ("110.69", 110.69),
    ("1098.30", 1098.30),   # >= 1000, confirmado real, sem milhar
    ("0.00", 0.0),
    ("1", 1.0),
])
def test_parse_brl_float_formato_real_ponto_decimal(raw, expected):
    assert mod._parse_brl_float(raw) == expected


def test_clean_numeric_formato_real_ponto_decimal():
    series, files, rows = _series(["110.69", "1098.30"])
    result = mod._clean_numeric(series, column="Subtotal do produto", brand="apice",
                                 source_files=files, source_rows=rows)
    assert list(result) == [110.69, 1098.30]


# --- bug histórico: separador de milhar BR virava NaN -> 0.0 silencioso ---

def test_comportamento_antigo_documentado_para_contraste():
    """Documenta o comportamento ANTERIOR (regex sem remoção de milhar +
    fillna(0.0)) só para contraste — não testa código de produção."""
    def _clean_numeric_antigo(series: pd.Series) -> pd.Series:
        return (
            series.astype(str)
            .str.replace(r"\s", "", regex=True)
            .str.replace(",", ".", regex=False)
            .pipe(pd.to_numeric, errors="coerce")
            .fillna(0.0)
        )

    old_result = _clean_numeric_antigo(pd.Series(["1.234,56"]))
    assert old_result.iloc[0] == 0.0  # bug antigo: silenciosamente zerado

    assert mod._parse_brl_float("1.234,56") == 1234.56  # comportamento novo


@pytest.mark.parametrize("raw, expected", [
    ("1.234,56", 1234.56),
    ("12.345,67", 12345.67),
])
def test_formato_br_com_milhar(raw, expected):
    assert mod._parse_brl_float(raw) == expected


def test_clean_numeric_formato_br_com_milhar():
    series, files, rows = _series(["1.234,56"])
    result = mod._clean_numeric(series, column="Subtotal do produto", brand="apice",
                                 source_files=files, source_rows=rows)
    assert list(result) == [1234.56]


@pytest.mark.parametrize("raw, expected", [
    ("1234,56", 1234.56),
    ("R$ 1.234,56", 1234.56),
    ("R$1.234,56", 1234.56),
    ("1.234,56\xa0", 1234.56),
    (" 1234.56 ", 1234.56),
])
def test_virgula_decimal_moeda_espacos_nbsp(raw, expected):
    assert mod._parse_brl_float(raw) == expected


# --- ausência legítima -> None / 0.0 na coluna, nunca erro ---

@pytest.mark.parametrize("raw", [None, "", "-", "N/A", "n/a", "NULL"])
def test_ausencia_legitima_retorna_none(raw):
    assert mod._parse_brl_float(raw) is None


def test_clean_numeric_ausencia_legitima_vira_zero_sem_erro():
    series, files, rows = _series([None, "", "-", "1098.30"])
    result = mod._clean_numeric(series, column="Subtotal do produto", brand="apice",
                                 source_files=files, source_rows=rows)
    assert list(result) == [0.0, 0.0, 0.0, 1098.30]


# --- valor não vazio inválido -> fail-fast, nunca 0.0 ---

def test_parse_brl_float_invalido_levanta_excecao():
    with pytest.raises(mod.ShopeeNumericParseError):
        mod._parse_brl_float("lixo_nao_numerico")


def test_clean_numeric_invalido_interrompe_antes_de_qualquer_zero():
    """Fail-fast: uma célula inválida no meio da série interrompe o
    processamento inteiro — nenhuma Series parcial com 0.0 é devolvida."""
    series, files, rows = _series(
        ["1098.30", "lixo_nao_numerico", "50.00"],
        files=["Order.all.A.xlsx", "Order.all.A.xlsx", "Order.all.A.xlsx"],
        rows=[2, 3, 4],
    )
    with pytest.raises(mod.ShopeeNumericParseError) as excinfo:
        mod._clean_numeric(series, column="Subtotal do produto", brand="apice",
                            source_files=files, source_rows=rows)
    message = str(excinfo.value)
    assert "apice" in message
    assert "Order.all.A.xlsx" in message
    assert "linha=3" in message
    assert "Subtotal do produto" in message
    assert "lixo_nao_numerico" not in message


# --- formato US rejeitado explicitamente, nunca convertido errado ---

@pytest.mark.parametrize("raw", ["1,234.56", "1,234,567.89", "12,345.6"])
def test_formato_us_rejeitado_explicitamente(raw):
    with pytest.raises(mod.ShopeeNumericParseError):
        mod._parse_brl_float(raw)


def test_formato_br_continua_funcionando_apos_rejeitar_us():
    assert mod._parse_brl_float("1.234,56") == 1234.56


# --- não finitos (NaN/Infinity), string e float nativo ---

@pytest.mark.parametrize("raw", ["NaN", "nan", "Infinity", "-Infinity", "inf", "-inf"])
def test_strings_nao_finitas_rejeitadas(raw):
    with pytest.raises(mod.ShopeeNumericParseError):
        mod._parse_brl_float(raw)


def test_float_nativo_infinity_rejeitado():
    """Ao contrário de NaN, Infinity nunca é usado por pandas como
    sentinela de ausência — só pode vir de dado genuinamente inválido."""
    with pytest.raises(mod.ShopeeNumericParseError):
        mod._parse_brl_float(float("inf"))
    with pytest.raises(mod.ShopeeNumericParseError):
        mod._parse_brl_float(float("-inf"))


def test_nan_nativo_e_ausencia_legitima_nao_erro():
    """Divergência deliberada do parser canônico (ver docstring de
    _parse_brl_float): confirmado empiricamente que
    `pd.read_excel(f, dtype=str)` representa tanto célula vazia quanto
    texto reconhecido como ausência ("NaN"/"NA"/"N/A"/"null") pelo MESMO
    sentinela float NaN, antes desta função ver o valor. Tratar NaN
    nativo como erro quebraria qualquer célula genuinamente vazia."""
    assert mod._parse_brl_float(math.nan) is None
    assert mod._parse_brl_float(float("nan")) is None


def test_clean_numeric_nan_nativo_de_pandas_vira_zero_sem_erro():
    """Reproduz o cenário real: pd.read_excel(..., dtype=str) já entrega
    célula vazia como float('nan') nativo (confirmado empiricamente nesta
    sessão, pandas 3.0.3) — a coluna resultante deve continuar float64
    com 0.0 na posição ausente, nunca levantar."""
    series = pd.Series([math.nan, "1098.30"])
    files, rows = pd.Series(["a.xlsx", "a.xlsx"]), pd.Series([2, 3])
    result = mod._clean_numeric(series, column="Subtotal do produto", brand="apice",
                                 source_files=files, source_rows=rows)
    assert list(result) == [0.0, 1098.30]
    assert result.dtype == "float64"


# --- dtype final sempre numérico, nunca object ---

def test_clean_numeric_dtype_sempre_float64():
    series, files, rows = _series(["110.69", "1098.30"])
    result = mod._clean_numeric(series, column="Subtotal do produto", brand="apice",
                                 source_files=files, source_rows=rows)
    assert result.dtype == "float64"


def test_clean_numeric_dtype_float64_mesmo_com_tudo_ausente():
    series, files, rows = _series([None, "", "-"])
    result = mod._clean_numeric(series, column="Subtotal do produto", brand="apice",
                                 source_files=files, source_rows=rows)
    assert result.dtype == "float64"


# --- sanitização: mensagem, traceback completo, __cause__, __context__ ---

_FICTITIOUS_SENSITIVE = [
    "nome_ficticio_joao_da_silva",
    "cpf_ficticio_123.456.789-00",
]


@pytest.mark.parametrize("raw", _FICTITIOUS_SENSITIVE + ["1,234.56"])
def test_parse_brl_float_nao_encadeia_cause_nem_context(raw):
    with pytest.raises(mod.ShopeeNumericParseError) as excinfo:
        mod._parse_brl_float(raw)
    exc = excinfo.value
    assert exc.__cause__ is None
    assert exc.__context__ is None


@pytest.mark.parametrize("raw", _FICTITIOUS_SENSITIVE)
def test_clean_numeric_traceback_completo_nao_vaza_valor_bruto(raw):
    series, files, rows = _series([raw])
    with pytest.raises(mod.ShopeeNumericParseError) as excinfo:
        mod._clean_numeric(series, column="Subtotal do produto", brand="apice",
                            source_files=files, source_rows=rows)
    exc = excinfo.value
    assert exc.__cause__ is None
    assert exc.__context__ is None
    formatted = "".join(traceback.format_exception(type(exc), exc, exc.__traceback__))
    assert raw not in formatted


# --- ausência de escrita/conexão quando o parsing falha ---

def test_parsing_invalido_nunca_referencia_conexao_ou_engine():
    """Confirma estruturalmente que _clean_numeric/_parse_brl_float nunca
    tocam engine/conexão — a falha de parsing acontece em memória, antes
    de qualquer possibilidade de abrir transação."""
    source = inspect.getsource(mod._clean_numeric) + inspect.getsource(mod._parse_brl_float)
    for forbidden in ("create_engine", "engine.begin", "conn.execute", "LOCAL_PG_URL"):
        assert forbidden not in source


def test_load_brand_propaga_erro_antes_de_qualquer_dropna_ou_agregacao(tmp_path, monkeypatch):
    """Fim a fim: um Subtotal inválido em _load_brand interrompe antes de
    chegar em _aggregate/main — nenhuma linha parcial é processada."""
    brand_dir = tmp_path / "apice"
    brand_dir.mkdir()
    header = [
        "Data de criação do pedido", "Nº de referência do SKU principal",
        "Nome do Produto", "Nome da variação", "Quantidade",
        "Subtotal do produto", "Status do pedido", "Nome de usuário (comprador)",
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    ws.append(["2026-01-05 10:00", "SKU1", "Produto A", None, "1", "lixo_nao_numerico", "Concluído", "u1"])
    wb.save(brand_dir / "Order.all.20260101_20260131.xlsx")

    monkeypatch.setattr(mod, "SHOPEE_ROOT", tmp_path)

    with pytest.raises(mod.ShopeeNumericParseError) as excinfo:
        mod._load_brand("apice")

    assert "lixo_nao_numerico" not in str(excinfo.value)
    assert "u1" not in str(excinfo.value)


# ---------------------------------------------------------------------------
# Quantidade (_parse_qty_int / _clean_int) — mesma classe de risco do
# subtotal: pd.to_numeric(...).fillna(0).astype(int) truncava/zerava
# silenciosamente. Reaproveita _parse_brl_float + validações de inteiro
# exato e não-negatividade.
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("raw, expected", [
    ("1", 1),
    ("2", 2),
    (1, 1),
    (2, 2),
])
def test_parse_qty_int_formatos_reais(raw, expected):
    assert mod._parse_qty_int(raw) == expected


@pytest.mark.parametrize("raw", [None, "", "-", "N/A", math.nan])
def test_parse_qty_int_ausencia_retorna_none(raw):
    assert mod._parse_qty_int(raw) is None


def test_clean_int_ausencia_vira_zero_sem_erro():
    series, files, rows = _series([None, "", math.nan, "2"])
    result = mod._clean_int(series, column="Quantidade", brand="apice",
                             source_files=files, source_rows=rows)
    assert list(result) == [0, 0, 0, 2]
    assert result.dtype == "int64"


def test_parse_qty_int_texto_invalido_levanta_excecao():
    with pytest.raises(mod.ShopeeNumericParseError):
        mod._parse_qty_int("lixo_nao_numerico")


def test_parse_qty_int_decimal_nao_inteiro_e_rejeitado_nunca_truncado():
    """"1.5" tem que falhar explicitamente — nunca virar 1 (truncamento
    silencioso) nem 2 (arredondamento silencioso)."""
    with pytest.raises(mod.ShopeeNumericParseError):
        mod._parse_qty_int("1.5")
    with pytest.raises(mod.ShopeeNumericParseError):
        mod._parse_qty_int("2,50")


def test_parse_qty_int_negativo_e_rejeitado():
    with pytest.raises(mod.ShopeeNumericParseError):
        mod._parse_qty_int("-1")
    with pytest.raises(mod.ShopeeNumericParseError):
        mod._parse_qty_int(-3)


def test_parse_qty_int_infinity_e_rejeitado():
    with pytest.raises(mod.ShopeeNumericParseError):
        mod._parse_qty_int("Infinity")
    with pytest.raises(mod.ShopeeNumericParseError):
        mod._parse_qty_int(float("inf"))


def test_parse_qty_int_formato_us_e_rejeitado():
    with pytest.raises(mod.ShopeeNumericParseError):
        mod._parse_qty_int("1,234.56")


def test_clean_int_invalido_interrompe_com_contexto_sanitizado():
    series, files, rows = _series(
        ["2", "1.5", "3"],
        files=["Order.all.A.xlsx"] * 3,
        rows=[2, 3, 4],
    )
    with pytest.raises(mod.ShopeeNumericParseError) as excinfo:
        mod._clean_int(series, column="Quantidade", brand="apice",
                        source_files=files, source_rows=rows)
    message = str(excinfo.value)
    assert "apice" in message
    assert "Order.all.A.xlsx" in message
    assert "linha=3" in message
    assert "Quantidade" in message
    assert "1.5" not in message


def test_clean_int_dtype_sempre_int64():
    series, files, rows = _series(["1", "2", "3"])
    result = mod._clean_int(series, column="Quantidade", brand="apice",
                             source_files=files, source_rows=rows)
    assert result.dtype == "int64"


@pytest.mark.parametrize("raw", _FICTITIOUS_SENSITIVE + ["1.5", "-1"])
def test_parse_qty_int_sanitizacao_completa(raw):
    with pytest.raises(mod.ShopeeNumericParseError) as excinfo:
        mod._parse_qty_int(raw)
    exc = excinfo.value
    assert exc.__cause__ is None
    assert exc.__context__ is None
    formatted = "".join(traceback.format_exception(type(exc), exc, exc.__traceback__))
    assert raw not in formatted


def test_load_brand_qty_invalida_propaga_antes_de_agregacao(tmp_path, monkeypatch):
    """Fim a fim: Quantidade inválida também interrompe _load_brand antes
    de qualquer dropna/agregação — mesmo padrão do Subtotal."""
    brand_dir = tmp_path / "apice"
    brand_dir.mkdir()
    header = [
        "Data de criação do pedido", "Nº de referência do SKU principal",
        "Nome do Produto", "Nome da variação", "Quantidade",
        "Subtotal do produto", "Status do pedido", "Nome de usuário (comprador)",
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    ws.append(["2026-01-05 10:00", "SKU1", "Produto A", None, "1.5", "100.00", "Concluído", "u1"])
    wb.save(brand_dir / "Order.all.20260101_20260131.xlsx")

    monkeypatch.setattr(mod, "SHOPEE_ROOT", tmp_path)

    with pytest.raises(mod.ShopeeNumericParseError) as excinfo:
        mod._load_brand("apice")

    assert "u1" not in str(excinfo.value)


# --- importação/execução no ambiente real de apps/api ---

def test_nunca_referencia_pipelines():
    """apps/api é empacotado e implantado de forma independente —
    `pipelines` nunca é instalado no seu .venv (confirmado empiricamente:
    `import pipelines` levanta ModuleNotFoundError com cwd=apps/api).
    Este teste protege contra alguém reintroduzir a IMPORT (não a mera
    menção em docstring/comentário, que este próprio arquivo já faz para
    explicar a decisão) sem reverificar essa premissa."""
    source = inspect.getsource(mod)
    assert not re.search(r"^\s*import pipelines\b", source, re.MULTILINE)
    assert not re.search(r"^\s*from pipelines\b", source, re.MULTILINE)
    # nenhuma linha de codigo real usa sys.path/importa sys (mencoes em
    # prosa de docstring, ex. explicando a decisao acima, nao comecam a
    # linha com esse padrao, entao nao disparam este check)
    assert not re.search(r"^\s*import sys\b", source, re.MULTILINE)
    assert not re.search(r"^\s*sys\.path", source, re.MULTILINE)


def test_modulo_importa_e_roda_no_ambiente_real_sem_pipelines():
    """O simples fato deste teste coletar e passar já prova que
    load_shopee_products importa e funciona no .venv real de apps/api
    (este arquivo de teste roda com
    apps/api/.venv/Scripts/python.exe -m pytest, cwd=repo root ou
    apps/api, sem pipelines instalado)."""
    assert callable(mod._clean_numeric)
    assert callable(mod._parse_brl_float)
    assert mod._parse_brl_float("1098.30") == 1098.30


# ---------------------------------------------------------------------------
# main() em duas fases: Fase A (memória/arquivos) nunca abre engine/
# conexão/DDL; Fase B (banco) só começa depois que TODAS as marcas foram
# validadas e agregadas com sucesso. Testes de FLUXO REAL com fakes —
# não meramente estruturais (inspect.getsource não prova comportamento
# em runtime, só ausência de padrões no texto-fonte).
# ---------------------------------------------------------------------------

def _write_valid_order_xlsx(path, buyer="u1"):
    header = [
        "Data de criação do pedido", "Nº de referência do SKU principal",
        "Nome do Produto", "Nome da variação", "Quantidade",
        "Subtotal do produto", "Status do pedido", "Nome de usuário (comprador)",
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    ws.append(["2026-01-05 10:00", "SKU1", "Produto A", None, "2", "100.00", "Concluído", buyer])
    wb.save(path)


class _FakeConnection:
    def __init__(self, calls):
        self._calls = calls

    def execute(self, stmt, params=None):
        self._calls.append((str(stmt), params))

    def __enter__(self):
        return self

    def __exit__(self, *exc):
        return False


class _FakeEngine:
    def __init__(self):
        self.calls: list[tuple[str, object]] = []

    def begin(self):
        return _FakeConnection(self.calls)


def test_main_nao_chama_create_engine_quando_load_brand_falha(tmp_path, monkeypatch):
    """Fluxo real (não estrutural): quando uma marca tem Subtotal inválido,
    _prepare_all_brands() levanta ShopeeNumericParseError e main() nunca
    chega a chamar _get_local_pg_url()/create_engine() — nem DDL nem
    qualquer conn.execute() acontece. Rastreamento via fakes que gravam
    toda chamada, não apenas asserts sobre o texto-fonte."""
    brand_dir = tmp_path / "apice"
    brand_dir.mkdir()
    header = [
        "Data de criação do pedido", "Nº de referência do SKU principal",
        "Nome do Produto", "Nome da variação", "Quantidade",
        "Subtotal do produto", "Status do pedido", "Nome de usuário (comprador)",
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    ws.append(["2026-01-05 10:00", "SKU1", "Produto A", None, "1", "lixo_nao_numerico", "Concluído", "u1"])
    wb.save(brand_dir / "Order.all.20260101_20260131.xlsx")

    monkeypatch.setattr(mod, "SHOPEE_ROOT", tmp_path)
    monkeypatch.setattr(mod, "BRANDS", ["apice"])

    pg_url_calls = []
    engine_calls = []
    monkeypatch.setattr(mod, "_get_local_pg_url", lambda: pg_url_calls.append(1))
    monkeypatch.setattr(mod, "create_engine", lambda *a, **k: engine_calls.append((a, k)))

    with pytest.raises(mod.ShopeeNumericParseError):
        mod.main()

    assert pg_url_calls == [], "_get_local_pg_url não pode ser chamado antes de todas as marcas serem validadas"
    assert engine_calls == [], "create_engine não pode ser chamado quando o parsing de qualquer marca falha"


def test_main_chama_create_engine_e_ddl_somente_apos_fase_a_completa(tmp_path, monkeypatch):
    """Caminho feliz, fluxo real: com todas as marcas válidas, main()
    chama _get_local_pg_url()/create_engine(), executa o DDL e só então
    grava os agregados — via engine/conexão FAKE (nenhum banco real)."""
    brand_dir = tmp_path / "apice"
    brand_dir.mkdir()
    _write_valid_order_xlsx(brand_dir / "Order.all.20260101_20260131.xlsx")

    monkeypatch.setattr(mod, "SHOPEE_ROOT", tmp_path)
    monkeypatch.setattr(mod, "BRANDS", ["apice"])
    monkeypatch.setattr(mod, "_get_local_pg_url", lambda: "postgresql://u:p@localhost/db")

    fake_engine = _FakeEngine()
    monkeypatch.setattr(mod, "create_engine", lambda *a, **k: fake_engine)

    mod.main()

    ddl_calls = [c for c in fake_engine.calls if "CREATE TABLE" in c[0]]
    upsert_calls = [c for c in fake_engine.calls if "INSERT INTO" in c[0]]
    assert len(ddl_calls) == 1, "DDL deve ser executado exatamente uma vez, na Fase B"
    assert len(upsert_calls) == 1, "agregado da marca válida deve ser gravado"
    # DDL sempre antes de qualquer INSERT (ordem das chamadas na fake engine)
    assert fake_engine.calls.index(ddl_calls[0]) < fake_engine.calls.index(upsert_calls[0])


def test_prepare_all_brands_descarta_dataframe_bruto_mantem_so_agregado(tmp_path, monkeypatch):
    """_prepare_all_brands devolve o agregado (poucas linhas por sku/mês),
    não o DataFrame bruto (uma linha por SKU por pedido) — confirma que a
    Fase A não acumula todos os brutos simultaneamente."""
    brand_dir = tmp_path / "apice"
    brand_dir.mkdir()
    _write_valid_order_xlsx(brand_dir / "Order.all.20260101_20260131.xlsx")

    monkeypatch.setattr(mod, "SHOPEE_ROOT", tmp_path)
    monkeypatch.setattr(mod, "BRANDS", ["apice"])

    prepared = mod._prepare_all_brands()

    assert len(prepared) == 1
    brand, agg = prepared[0]
    assert brand == "apice"
    # agregado tem as colunas de _aggregate, nunca as colunas brutas de
    # _load_brand (ex.: "status" some, "completed_orders" aparece)
    assert "completed_orders" in agg.columns
    assert "status" not in agg.columns
