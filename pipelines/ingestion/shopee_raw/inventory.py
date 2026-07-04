"""
Inventário e leitura read-only dos exports locais da Shopee.

Este módulo nunca escreve em disco nem em banco de dados. Ele só lê
`shopee/{brand}/...` e devolve estruturas Python descrevendo o que existe
(inventário) ou o que seria inserido (dry-run), preservando cada linha
física do jeito que a Shopee exportou.
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Optional

import openpyxl

from pipelines.connectors.shopee.connector import BRANDS_IN_SCOPE as OFFICIAL_BRANDS
from pipelines.ingestion.shopee_raw.hashing import row_sha256, sha256_file, sha256_text

SOURCE_ORDERS = "orders"
SOURCE_SHOP_STATS = "shop_stats"
SOURCE_ADS = "ads"
SOURCE_UNKNOWN = "unknown"

_ORDERS_HEADER_ROW = 0
_SHOPSTATS_HEADER_ROW = 3
_SHOPSTATS_TOTAL_ROW = 1
_SHOPSTATS_MIN_ROWS = 5

_PART_RE = re.compile(r"_part_(\d+)_of_(\d+)", re.IGNORECASE)
_DATE8_UNDERSCORE_RE = re.compile(r"(\d{8})_(\d{8})")
_DATE8_DASH_RE = re.compile(r"(\d{8})-(\d{8})")


class SourceReadError(Exception):
    """Erro ao ler o conteúdo de um arquivo de export Shopee."""


def detect_source_type(filename: str) -> str:
    if filename.startswith("Order.all") and filename.endswith(".xlsx"):
        return SOURCE_ORDERS
    if ".shopee-shop-stats." in filename and filename.endswith(".xlsx"):
        return SOURCE_SHOP_STATS
    if filename.startswith("Dados") and filename.endswith(".csv"):
        return SOURCE_ADS
    return SOURCE_UNKNOWN


def parse_filename_part(filename: str) -> tuple[str, Optional[int], Optional[int]]:
    """'Order.all.X_part_2_of_6.xlsx' -> ('Order.all.X.xlsx', 2, 6)."""
    m = _PART_RE.search(filename)
    if not m:
        return filename, None, None
    part_index, part_total = int(m.group(1)), int(m.group(2))
    group = filename[: m.start()] + filename[m.end():]
    return group, part_index, part_total


def parse_filename_date_range(filename: str) -> tuple[Optional[date], Optional[date]]:
    """Extrai um período do nome do arquivo, somente para diagnóstico."""
    for rx in (_DATE8_UNDERSCORE_RE, _DATE8_DASH_RE):
        m = rx.search(filename)
        if m:
            try:
                d_from = datetime.strptime(m.group(1), "%Y%m%d").date()
                d_to = datetime.strptime(m.group(2), "%Y%m%d").date()
                return d_from, d_to
            except ValueError:
                continue
    return None, None


@dataclass
class ParsedRow:
    source_row_number: int
    raw_payload: dict
    row_sha256: str


@dataclass
class RowReject:
    source_row_number: int
    reason: str


@dataclass
class RowReadResult:
    headers: list[str]
    header_row_index: Optional[int]
    sheet_name: Optional[str]
    rows: list[ParsedRow]
    rejects: list[RowReject]
    encoding: Optional[str] = None
    delimiter: Optional[str] = None


def _is_blank_row(row: tuple) -> bool:
    return all(v is None or (isinstance(v, str) and v.strip() == "") for v in row)


def _row_to_payload(headers: list[str], row: tuple) -> dict:
    payload = {}
    for i, header in enumerate(headers):
        if header is None:
            continue
        key = header if header not in payload else f"{header}__col{i}"
        payload[key] = row[i] if i < len(row) else None
    return payload


def read_orders_file(path: Path) -> RowReadResult:
    """Lê Order.all*.xlsx. Grão: uma linha física de SKU por pedido."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:  # noqa: BLE001 - queremos capturar qualquer falha de leitura
        raise SourceReadError(f"falha ao abrir xlsx: {exc}") from exc

    try:
        ws = wb.active
        sheet_name = ws.title
        raw_rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not raw_rows:
        return RowReadResult(headers=[], header_row_index=None, sheet_name=sheet_name, rows=[], rejects=[])

    header = list(raw_rows[_ORDERS_HEADER_ROW])
    rows: list[ParsedRow] = []
    rejects: list[RowReject] = []
    for i, raw_row in enumerate(raw_rows[_ORDERS_HEADER_ROW + 1:], start=_ORDERS_HEADER_ROW + 1):
        if _is_blank_row(raw_row):
            rejects.append(RowReject(source_row_number=i, reason="linha completamente vazia"))
            continue
        payload = _row_to_payload(header, raw_row)
        rows.append(ParsedRow(source_row_number=i, raw_payload=payload, row_sha256=row_sha256(payload)))

    return RowReadResult(
        headers=header, header_row_index=_ORDERS_HEADER_ROW, sheet_name=sheet_name, rows=rows, rejects=rejects
    )


def read_shop_stats_file(path: Path) -> RowReadResult:
    """Lê shop-stats xlsx. Grão: uma linha física (total do período OU dia)."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise SourceReadError(f"falha ao abrir xlsx: {exc}") from exc

    try:
        ws = wb.active
        sheet_name = ws.title
        raw_rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if len(raw_rows) < _SHOPSTATS_MIN_ROWS:
        raise SourceReadError(
            f"template shop-stats esperado tem >= {_SHOPSTATS_MIN_ROWS} linhas; encontrado {len(raw_rows)}"
        )

    header = list(raw_rows[_SHOPSTATS_HEADER_ROW])
    rows: list[ParsedRow] = []
    rejects: list[RowReject] = []

    total_row = raw_rows[_SHOPSTATS_TOTAL_ROW]
    if _is_blank_row(total_row):
        rejects.append(RowReject(source_row_number=_SHOPSTATS_TOTAL_ROW, reason="linha de total do período vazia"))
    else:
        payload = _row_to_payload(header, total_row)
        rows.append(
            ParsedRow(
                source_row_number=_SHOPSTATS_TOTAL_ROW, raw_payload=payload, row_sha256=row_sha256(payload)
            )
        )

    for i, raw_row in enumerate(raw_rows[_SHOPSTATS_HEADER_ROW + 1:], start=_SHOPSTATS_HEADER_ROW + 1):
        if _is_blank_row(raw_row):
            rejects.append(RowReject(source_row_number=i, reason="linha completamente vazia"))
            continue
        payload = _row_to_payload(header, raw_row)
        rows.append(ParsedRow(source_row_number=i, raw_payload=payload, row_sha256=row_sha256(payload)))

    return RowReadResult(
        headers=header, header_row_index=_SHOPSTATS_HEADER_ROW, sheet_name=sheet_name, rows=rows, rejects=rejects
    )


def _decode_ads_csv(path: Path) -> tuple[list[str], str]:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            text = raw.decode(encoding)
            return text.splitlines(), encoding
        except UnicodeDecodeError:
            continue
    raise SourceReadError("não foi possível decodificar o CSV com utf-8-sig/cp1252/latin-1")


def read_ads_file(path: Path) -> RowReadResult:
    """Lê CSV de ads. Grão: uma linha física por anúncio."""
    lines, encoding = _decode_ads_csv(path)

    header_line_idx = None
    for i, line in enumerate(lines):
        if line.startswith("#,"):
            header_line_idx = i
            break

    if header_line_idx is None:
        raise SourceReadError("header de colunas ('#,...') não encontrado no CSV")

    delimiter = ","
    reader = csv.DictReader(io.StringIO("\n".join(lines[header_line_idx:])), delimiter=delimiter)
    header = list(reader.fieldnames or [])

    rows: list[ParsedRow] = []
    rejects: list[RowReject] = []
    for i, raw_row in enumerate(reader, start=header_line_idx + 1):
        values = list(raw_row.values())
        if all(v is None or str(v).strip() == "" for v in values):
            rejects.append(RowReject(source_row_number=i, reason="linha completamente vazia"))
            continue
        payload = dict(raw_row)
        rows.append(ParsedRow(source_row_number=i, raw_payload=payload, row_sha256=row_sha256(payload)))

    return RowReadResult(
        headers=header,
        header_row_index=header_line_idx,
        sheet_name=None,
        rows=rows,
        rejects=rejects,
        encoding=encoding,
        delimiter=delimiter,
    )


_READERS = {
    SOURCE_ORDERS: read_orders_file,
    SOURCE_SHOP_STATS: read_shop_stats_file,
    SOURCE_ADS: read_ads_file,
}


def read_source_file(path: Path, source_type: str) -> RowReadResult:
    reader = _READERS.get(source_type)
    if reader is None:
        raise SourceReadError(f"source_type desconhecido, sem leitor: {source_type!r}")
    return reader(path)


def compute_schema_fingerprint(headers: list[str]) -> str:
    normalized = "|".join("" if h is None else str(h) for h in headers)
    return sha256_text(normalized)


@dataclass
class FileInventoryRecord:
    relative_path: str
    brand: str
    brand_known: bool
    source_type: str
    extension: str
    size_bytes: int
    file_sha256: Optional[str]
    source_modified_at: str
    sheet_name: Optional[str] = None
    header_row_index: Optional[int] = None
    headers: Optional[list[str]] = None
    source_row_count: Optional[int] = None
    rejected_row_count: int = 0
    schema_fingerprint: Optional[str] = None
    encoding: Optional[str] = None
    delimiter: Optional[str] = None
    is_empty: bool = False
    is_readable: bool = True
    error_message: Optional[str] = None
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    part_group: Optional[str] = None
    part_index: Optional[int] = None
    part_total: Optional[int] = None


def scan_file(path: Path, base_dir: Path, official_brands: tuple[str, ...] = OFFICIAL_BRANDS) -> FileInventoryRecord:
    relative_path = path.relative_to(base_dir).as_posix()
    brand = path.relative_to(base_dir).parts[0]
    brand_known = brand in official_brands
    filename = path.name
    source_type = detect_source_type(filename)
    extension = path.suffix.lower()

    size_bytes = path.stat().st_size
    mtime = datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).isoformat()
    part_group, part_index, part_total = parse_filename_part(filename)
    date_from, date_to = parse_filename_date_range(filename)

    record = FileInventoryRecord(
        relative_path=relative_path,
        brand=brand,
        brand_known=brand_known,
        source_type=source_type,
        extension=extension,
        size_bytes=size_bytes,
        file_sha256=None,
        source_modified_at=mtime,
        is_empty=(size_bytes == 0),
        part_group=part_group,
        part_index=part_index,
        part_total=part_total,
        date_from=date_from.isoformat() if date_from else None,
        date_to=date_to.isoformat() if date_to else None,
    )

    if record.is_empty:
        record.is_readable = False
        record.error_message = "arquivo vazio (0 bytes)"
        return record

    try:
        record.file_sha256 = sha256_file(path)
    except OSError as exc:
        record.is_readable = False
        record.error_message = f"falha ao ler arquivo para hash: {exc}"
        return record

    if source_type == SOURCE_UNKNOWN:
        # Arquivo aparece no relatório, mas não tentamos parsear conteúdo de
        # tipo desconhecido — evita interpretar formato não mapeado.
        return record

    try:
        result = read_source_file(path, source_type)
    except SourceReadError as exc:
        record.is_readable = False
        record.error_message = str(exc)
        return record

    record.sheet_name = result.sheet_name
    record.header_row_index = result.header_row_index
    record.headers = result.headers
    record.source_row_count = len(result.rows)
    record.rejected_row_count = len(result.rejects)
    record.encoding = result.encoding
    record.delimiter = result.delimiter
    record.schema_fingerprint = compute_schema_fingerprint(result.headers) if result.headers else None
    return record


def scan_directory(base_dir: Path, official_brands: tuple[str, ...] = OFFICIAL_BRANDS) -> list[FileInventoryRecord]:
    if not base_dir.exists():
        raise FileNotFoundError(f"diretório não encontrado: {base_dir}")

    records: list[FileInventoryRecord] = []
    for entry in sorted(base_dir.iterdir()):
        if not entry.is_dir():
            continue
        for path in sorted(entry.rglob("*")):
            if not path.is_file():
                continue
            if path.name.startswith("."):
                continue
            records.append(scan_file(path, base_dir, official_brands))
    return records


def find_duplicate_files(records: list[FileInventoryRecord]) -> dict[str, list[str]]:
    by_hash: dict[str, list[str]] = {}
    for r in records:
        if r.file_sha256 is None:
            continue
        by_hash.setdefault(r.file_sha256, []).append(r.relative_path)
    return {h: paths for h, paths in by_hash.items() if len(paths) > 1}


def find_schema_drift(records: list[FileInventoryRecord]) -> dict[str, dict[str, list[str]]]:
    """Por source_type, agrupa fingerprints distintos -> arquivos que os usam.

    Mais de um fingerprint no mesmo source_type indica diferença de template
    (ex: entre marcas, ou dentro da mesma marca ao longo do tempo).
    """
    by_source: dict[str, dict[str, list[str]]] = {}
    for r in records:
        if not r.schema_fingerprint:
            continue
        by_source.setdefault(r.source_type, {}).setdefault(r.schema_fingerprint, []).append(r.relative_path)
    return {source: fps for source, fps in by_source.items() if len(fps) > 1}


def find_overlapping_exports(records: list[FileInventoryRecord]) -> list[dict]:
    """Detecta grupos de arquivos (por brand+source_type) com períodos que se
    sobrepõem, comparando por part_group (arquivos partidos do mesmo export
    não contam como sobreposição entre si)."""
    groups: dict[tuple[str, str], dict[str, tuple[date, date]]] = {}
    for r in records:
        if r.source_type not in (SOURCE_ORDERS, SOURCE_SHOP_STATS) or not r.date_from or not r.date_to:
            continue
        key = (r.brand, r.source_type)
        d_from = date.fromisoformat(r.date_from)
        d_to = date.fromisoformat(r.date_to)
        existing = groups.setdefault(key, {})
        if r.part_group not in existing:
            existing[r.part_group] = (d_from, d_to)
        else:
            prev_from, prev_to = existing[r.part_group]
            existing[r.part_group] = (min(prev_from, d_from), max(prev_to, d_to))

    overlaps = []
    for (brand, source_type), part_ranges in groups.items():
        items = list(part_ranges.items())
        for i in range(len(items)):
            for j in range(i + 1, len(items)):
                (group_a, (a_from, a_to)) = items[i]
                (group_b, (b_from, b_to)) = items[j]
                if a_from <= b_to and b_from <= a_to:
                    overlaps.append(
                        {
                            "brand": brand,
                            "source_type": source_type,
                            "group_a": group_a,
                            "range_a": (a_from.isoformat(), a_to.isoformat()),
                            "group_b": group_b,
                            "range_b": (b_from.isoformat(), b_to.isoformat()),
                        }
                    )
    return overlaps


def build_summary(records: list[FileInventoryRecord]) -> dict:
    total_size = sum(r.size_bytes for r in records)
    total_rows = sum(r.source_row_count or 0 for r in records if r.is_readable)
    by_source_type: dict[str, int] = {}
    by_brand: dict[str, int] = {}
    for r in records:
        by_source_type[r.source_type] = by_source_type.get(r.source_type, 0) + 1
        by_brand[r.brand] = by_brand.get(r.brand, 0) + 1

    unreadable = [r.relative_path for r in records if not r.is_readable]
    unknown = [r.relative_path for r in records if r.source_type == SOURCE_UNKNOWN or not r.brand_known]

    return {
        "total_files": len(records),
        "total_size_bytes": total_size,
        "total_readable_rows": total_rows,
        "by_source_type": by_source_type,
        "by_brand": by_brand,
        "unreadable_files": unreadable,
        "unknown_or_unlisted_files": unknown,
        "duplicate_files_by_sha256": find_duplicate_files(records),
        "schema_drift_by_source_type": find_schema_drift(records),
        "overlapping_exports": find_overlapping_exports(records),
    }
