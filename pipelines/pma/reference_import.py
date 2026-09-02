"""Gate PMA-1A — importador do snapshot de preco sugerido de revenda (PDV).

CLI SEPARADO DO SYNC DE PRECOS, DE PROPOSITO
--------------------------------------------
Sao dois fatos de natureza diferente. O sync de precos reescreve uma janela e e'
idempotente; este importador ACRESCENTA um snapshot e nunca sobrescreve. Um unico
CLI com dois modos acabaria compartilhando lock, janela e reconciliacao entre
semanticas incompativeis.

O QUE E' LIDO — E O QUE NUNCA E'
-------------------------------
Le SOMENTE a tabela de produtos da aba "Geral*", a partir da linha 33 (cabecalho
na 32). As linhas 1-31 sao o cabecalho do PEDIDO e contem o BLOCO CADASTRAL:
razao social, CNPJ, I.E., CEP, endereco, bairro, cidade, estado, telefone,
e-mail. Esse bloco nao e' lido, nao e' logado, nao e' escrito no snapshot e nao
tem coluna na tabela de destino.

Tres travas independentes, porque uma so falharia em silencio se o layout mudar:
  1. ESTRUTURAL — a varredura comeca em `FIRST_DATA_ROW`; nada acima e' tocado;
  2. POR NOME    — as colunas sao resolvidas por cabecalho de uma allowlist
                   (`COLUMN_ALIASES`); coluna fora dela nunca e' lida;
  3. POR FORMA   — todo texto que vai ao snapshot passa por
                   `assert_no_pii_shaped_value` (CNPJ/CPF/e-mail/telefone/CEP).

NADA VERSIONADO
---------------
O snapshot sanitizado e' escrito FORA do repositorio (`--out`, obrigatorio, e
recusado se apontar para dentro da arvore de trabalho). XLSX, CSV com preco e
dado bruto nunca entram no git.

SEM `--apply`, NAO ESCREVE
--------------------------
Sem a flag o CLI le, valida, classifica, grava o snapshot local e relata. Com
`--apply` ele escreve o snapshot em `marts.fact_suggested_price_reference_snapshot`
no Neon, dentro de UMA transacao, sob advisory lock proprio, e RECUSA se o
`snapshot_id` ja existir — append-only nao sobrescreve.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import unicodedata
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path

from pipelines.pma.reference_contract import (
    COLUMN_ALIASES,
    FIRST_DATA_ROW,
    HEADER_ROW,
    IGNORED_COLUMNS,
    QUALITY_MISSING_PRICE,
    QUALITY_STATUSES,
    REFERENCE_BRANDS,
    REFERENCE_TYPE,
    SHEET_PREFIX,
    VALIDITY_STATUS,
    POLICY_STATUS,
    ReferenceContractError,
    ReferenceRow,
    assert_no_pii_shaped_value,
    assert_payload_has_no_forbidden_terms,
    classify_gtin,
    classify_quality,
    normalize_header,
    normalize_sku,
    notes_to_text,
    reference_row_id,
    snapshot_id_for,
)

TARGET_TABLE = "marts.fact_suggested_price_reference_snapshot"
STAGING_TABLE = "stg_pma_reference_publish"

#: Advisory lock proprio deste importador. Chave distinta da do sync de precos:
#: as duas rotinas escrevem tabelas diferentes e nao devem se bloquear.
ADVISORY_LOCK_KEY = 913_120_013

TARGET_STATEMENT_TIMEOUT = "300s"
CONNECT_TIMEOUT_SECONDS = 15
INSERT_PAGE_SIZE = 500
LOCK_TIMEOUT = "30s"

BUSINESS_COLUMNS = (
    "snapshot_id", "reference_row_id", "captured_at", "brand",
    "source_sku", "source_gtin", "product_name",
    "wholesale_amount", "suggested_retail_amount",
    "reference_type", "validity_status", "quality_status", "quality_notes",
    "source_file_hash",
)
AUDIT_COLUMNS = ("source_run_id",)

_RUN_ID_RE = re.compile(r"[^A-Za-z0-9_:-]")
_BRAND_FROM_NAME_RE = re.compile(r"[a-z0-9]+")


def _strip_accents_lower(texto: str) -> str:
    """Minusculo sem acento, PRESERVANDO os separadores.

    `reference_contract.normalize_brand` colapsa espacos — serve para comparar um
    valor de celula com a allowlist, e nao para tokenizar nome de arquivo.
    Reusa-la aqui transformaria "Tabela de Preco Yenzah" em um unico token.
    """
    nfkd = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in nfkd if not unicodedata.combining(c)).lower()


class ReferenceImportError(RuntimeError):
    """Falha do importador. Mensagem sem valor de celula e sem topologia."""


def sanitize_run_id(raw: str) -> str:
    return _RUN_ID_RE.sub("_", raw)[:64]


def default_run_id(now: datetime | None = None) -> str:
    stamp = (now or datetime.now(timezone.utc)).strftime("%Y%m%dT%H%M%SZ")
    return sanitize_run_id(f"pma_reference_import:{stamp}")


def file_sha256(caminho: Path) -> str:
    """Hash do arquivo-fonte. Prova qual conteudo gerou o snapshot."""
    h = hashlib.sha256()
    with caminho.open("rb") as fh:
        for bloco in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(bloco)
    return h.hexdigest()


def brand_from_filename(caminho: Path) -> str:
    """Marca a partir do nome do arquivo, validada contra a allowlist.

    Nao e' adivinhacao: o nome e' TOKENIZADO (acentos removidos, minusculo,
    quebra em qualquer caractere nao alfanumerico) e cada token e' comparado por
    IGUALDADE contra `REFERENCE_BRANDS`. Assim "Tabela de Preco Yenzah.xlsx" ->
    tokens {tabela, de, preco, yenzah} -> `yenzah`.

    Comparacao por substring seria mais permissiva e tambem mais perigosa: um
    nome como "kokeshi_e_yenzah" casaria duas marcas. Zero ou duas marcas
    encontradas e' RECUSA — importar sob a marca errada quebraria o escopo do
    match, que e' a garantia central deste gate.
    """
    tokens = set(_BRAND_FROM_NAME_RE.findall(_strip_accents_lower(caminho.stem)))
    achadas = [b for b in REFERENCE_BRANDS if b in tokens]
    if len(achadas) != 1:
        raise ReferenceImportError(
            f"nao foi possivel determinar a marca do arquivo {caminho.name!r} de "
            f"forma inequivoca (candidatas: {achadas or 'nenhuma'}). Renomeie o "
            f"arquivo para conter exatamente uma de: {', '.join(REFERENCE_BRANDS)}."
        )
    return achadas[0]


def _to_decimal(valor: object, campo: str, linha: int) -> Decimal | None:
    """Numero -> Decimal(2). `None` quando ausente. NUNCA devolve zero por ausencia."""
    if valor is None or (isinstance(valor, str) and not valor.strip()):
        return None
    try:
        d = Decimal(str(valor))
    except (InvalidOperation, ValueError):
        raise ReferenceImportError(
            f"valor nao numerico no campo {campo!r} da linha {linha}: recusado."
        )
    if d.is_nan():
        raise ReferenceImportError(
            f"NaN no campo {campo!r} da linha {linha}: NaN nao e' preco."
        )
    if d <= 0:
        # Zero e negativo nao sao referencia. Devolver None faria a linha cair em
        # `missing_suggested_price`, que e' o diagnostico honesto.
        return None
    return d.quantize(Decimal("0.01"))


def _resolve_columns(sheet, header_row: int) -> dict[str, int]:
    """Mapeia campo do contrato -> indice de coluna, por NOME de cabecalho.

    Resolucao posicional seria um defeito: Apice/Barbours/Kokeshi tem cinco
    colunas extra (Volumetria, Quantidade por caixa, CURVA, DESCRICAO) entre
    `nome cadastro` e `preco atacado`, e Rituaria/Yenzah nao tem nenhuma.
    """
    cabecalhos: dict[str, int] = {}
    for col in range(1, (sheet.max_column or 0) + 1):
        nome = normalize_header(sheet.cell(header_row, col).value)
        if nome and nome not in cabecalhos:
            cabecalhos[nome] = col

    resolvidas: dict[str, int] = {}
    for campo, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in cabecalhos:
                resolvidas[campo] = cabecalhos[alias]
                break

    obrigatorias = ("source_sku", "product_name", "suggested_retail_amount")
    faltando = [c for c in obrigatorias if c not in resolvidas]
    if faltando:
        raise ReferenceImportError(
            f"cabecalho da linha {header_row} nao tem as colunas obrigatorias "
            f"{faltando}: o layout da planilha mudou e a importacao para em vez "
            f"de adivinhar posicoes."
        )
    return resolvidas


def read_reference_file(caminho: Path) -> tuple[list[ReferenceRow], dict]:
    """Le UM arquivo e devolve linhas sanitizadas + relatorio de exclusoes.

    `openpyxl` e' importado AQUI, nao no topo do modulo: manter o import local
    deixa o modulo carregavel (e o contrato inspecionavel) num ambiente sem a
    dependencia, e nada e' aberto na importacao.
    """
    import openpyxl  # noqa: PLC0415 — ver docstring

    marca = brand_from_filename(caminho)
    hash_arquivo = file_sha256(caminho)
    # `data_only=True`: queremos o VALOR calculado do PDV, nao a formula. A
    # Rituaria tem `=ROUND(atacado*1,6;2)` e as outras quatro tem valor digitado.
    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=False)
    try:
        abas = [ws for ws in wb.worksheets
                if normalize_header(ws.title).startswith(SHEET_PREFIX)]
        if len(abas) != 1:
            raise ReferenceImportError(
                f"{caminho.name}: esperava exatamente uma aba comecando por "
                f"{SHEET_PREFIX!r}, encontrei {len(abas)}."
            )
        sheet = abas[0]
        colunas = _resolve_columns(sheet, HEADER_ROW)

        linhas: list[ReferenceRow] = []
        exclusoes: dict[str, int] = {}

        def excluir(motivo: str) -> None:
            exclusoes[motivo] = exclusoes.get(motivo, 0) + 1

        for r in range(FIRST_DATA_ROW, (sheet.max_row or 0) + 1):
            sku = normalize_sku(sheet.cell(r, colunas["source_sku"]).value)
            if sku is None:
                excluir("linha sem SKU (fora da tabela de produtos)")
                continue

            nome_bruto = sheet.cell(r, colunas["product_name"]).value
            nome = "" if nome_bruto is None else str(nome_bruto).strip()
            if not nome:
                excluir("produto sem nome")
                continue
            # Trava 3 — forma. Se o layout mudar e uma coluna cadastral cair no
            # lugar do nome do produto, para aqui em vez de vazar.
            assert_no_pii_shaped_value(nome, "product_name")
            assert_no_pii_shaped_value(sku, "source_sku")

            gtin_col = colunas.get("source_gtin")
            gtin_bruto = sheet.cell(r, gtin_col).value if gtin_col else None
            gtin = classify_gtin(gtin_bruto)

            atacado = _to_decimal(
                sheet.cell(r, colunas["wholesale_amount"]).value
                if "wholesale_amount" in colunas else None,
                "wholesale_amount", r,
            )
            pdv = _to_decimal(
                sheet.cell(r, colunas["suggested_retail_amount"]).value,
                "suggested_retail_amount", r,
            )

            notas: list[str] = []
            if gtin.note:
                notas.append(gtin.note)

            linhas.append(ReferenceRow(
                brand=marca,
                reference_row_id=reference_row_id(
                    marca, sku, str(gtin_bruto or ""), r
                ),
                source_sku=sku,
                source_gtin=gtin.value,
                product_name=nome,
                wholesale_amount=atacado,
                suggested_retail_amount=pdv,
                source_row_number=r,
                quality_notes=notas,
            ))

        relatorio = {
            "file": caminho.name,
            "brand": marca,
            "source_file_hash": hash_arquivo,
            "sheet": sheet.title,
            "header_row": HEADER_ROW,
            "first_data_row": FIRST_DATA_ROW,
            "rows_read": len(linhas),
            "exclusions": exclusoes,
            "ignored_columns": list(IGNORED_COLUMNS),
        }
        return linhas, relatorio
    finally:
        wb.close()


def build_snapshot(arquivos: list[Path], captured_at: datetime,
                   run_id: str) -> dict:
    """Le todos os arquivos, classifica ambiguidade e monta o snapshot completo.

    A classificacao roda DEPOIS de ler tudo, porque ambiguidade e' propriedade do
    conjunto: um SKU so e' ambiguo porque outra linha da MESMA marca o repete.
    """
    todas: list[ReferenceRow] = []
    relatorios: list[dict] = []
    hash_por_marca: dict[str, str] = {}
    for caminho in arquivos:
        linhas, rel = read_reference_file(caminho)
        todas.extend(linhas)
        relatorios.append(rel)
        hash_por_marca[rel["brand"]] = rel["source_file_hash"]

    duplicadas = _duplicate_row_ids(todas)
    if duplicadas:
        raise ReferenceImportError(
            f"{duplicadas} reference_row_id repetidos: a identidade deterministica "
            f"colidiu, o que quebraria a PK do snapshot. Importacao abortada."
        )

    classify_quality(todas)
    snapshot_id = snapshot_id_for(captured_at)

    registros = []
    for r in todas:
        registros.append({
            "snapshot_id": snapshot_id,
            "reference_row_id": r.reference_row_id,
            "captured_at": captured_at,
            "brand": r.brand,
            "source_sku": r.source_sku,
            "source_gtin": r.source_gtin,
            "product_name": r.product_name,
            "wholesale_amount": r.wholesale_amount,
            "suggested_retail_amount": r.suggested_retail_amount,
            "reference_type": REFERENCE_TYPE,
            "validity_status": VALIDITY_STATUS,
            "quality_status": r.quality_status,
            "quality_notes": notes_to_text(r.quality_notes),
            "source_file_hash": hash_por_marca[r.brand],
            "source_run_id": run_id,
        })

    _assert_snapshot_contract(registros)

    por_marca: dict[str, dict] = {}
    for reg in registros:
        alvo = por_marca.setdefault(
            reg["brand"], {"rows": 0, "quality": {}, "with_gtin": 0, "usable": 0}
        )
        alvo["rows"] += 1
        alvo["quality"][reg["quality_status"]] = (
            alvo["quality"].get(reg["quality_status"], 0) + 1
        )
        if reg["source_gtin"]:
            alvo["with_gtin"] += 1
        if reg["suggested_retail_amount"] is not None:
            alvo["usable"] += 1

    return {
        "snapshot_id": snapshot_id,
        "captured_at": captured_at,
        "run_id": run_id,
        "reference_type": REFERENCE_TYPE,
        "validity_status": VALIDITY_STATUS,
        "policy_status": POLICY_STATUS,
        "rows": registros,
        "files": relatorios,
        "by_brand": por_marca,
        "total_rows": len(registros),
    }


def _duplicate_row_ids(linhas: list[ReferenceRow]) -> int:
    vistos: set[str] = set()
    repetidos = 0
    for r in linhas:
        if r.reference_row_id in vistos:
            repetidos += 1
        vistos.add(r.reference_row_id)
    return repetidos


def _assert_snapshot_contract(registros: list[dict]) -> None:
    """Ultima fronteira antes de qualquer escrita. Espelha os CHECK do DDL."""
    for reg in registros:
        if reg["reference_type"] != REFERENCE_TYPE:
            raise ReferenceImportError(
                f"reference_type invalido: {reg['reference_type']!r}. O unico valor "
                f"aceito e' {REFERENCE_TYPE!r} — PDV nao e' PMA."
            )
        if reg["validity_status"] != VALIDITY_STATUS:
            raise ReferenceImportError(
                "validity_status precisa permanecer 'missing': as planilhas nao "
                "tem vigencia e nenhuma data pode ser inventada."
            )
        if reg["quality_status"] not in QUALITY_STATUSES:
            raise ReferenceImportError(
                f"quality_status desconhecido: {reg['quality_status']!r}."
            )
        if reg["brand"] not in REFERENCE_BRANDS:
            raise ReferenceImportError(f"marca fora da allowlist: {reg['brand']!r}.")
        gtin = reg["source_gtin"]
        if gtin is not None and (not gtin.isdigit() or len(gtin) not in (8, 12, 13)):
            raise ReferenceImportError(
                "source_gtin invalido no snapshot: somente EAN de consumidor "
                "(8/12/13 digitos) pode ocupar essa coluna. Codigo de 14 digitos "
                "e' DUN e precisa ficar nulo."
            )
        # BICONDICIONAL, espelhando `ck_fsprs_suggested_retail_amount` do DDL:
        # `missing_suggested_price` <=> valor NULO.  (PMA-1A-R, F1)
        #
        # A versao anterior deste importador produzia linhas com
        # `quality_status='missing_suggested_price'` e `suggested_retail_amount
        # = None` contra um DDL que declarava a coluna NOT NULL — a publicacao
        # teria falhado no primeiro `--apply` com preco em branco na planilha.
        # A coluna agora e' anulavel e a regra e' verificada nos DOIS sentidos,
        # aqui e no banco.
        pdv = reg["suggested_retail_amount"]
        ausente = reg["quality_status"] == QUALITY_MISSING_PRICE
        if ausente and pdv is not None:
            raise ReferenceImportError(
                "quality_status=missing_suggested_price exige "
                "suggested_retail_amount NULO: um valor aqui mentiria sobre a "
                "planilha."
            )
        if not ausente and pdv is None:
            raise ReferenceImportError(
                "suggested_retail_amount NULO exige "
                "quality_status=missing_suggested_price: sem isso a linha viraria "
                "referencia fantasma."
            )
        if pdv is not None and Decimal(str(pdv)) <= 0:
            raise ReferenceImportError(
                "suggested_retail_amount precisa ser > 0: referencia igual a zero "
                "nao e' referencia. Ausencia se registra como NULO."
            )

        # `quality_notes` e' texto que NOS geramos e que chega ao payload da
        # tela. Passa pelo guard de vocabulario; `product_name` NAO passa, de
        # proposito: e' texto do fornecedor, e reprovar um nome de produto por
        # conter uma palavra generica seria falso positivo caro.
        notas = reg.get("quality_notes")
        if notas:
            assert_payload_has_no_forbidden_terms(notas, "quality_notes")


# ---------------------------------------------------------------------------
# Snapshot local — fora do repositorio, obrigatoriamente
# ---------------------------------------------------------------------------

def assert_outside_repo(destino: Path) -> Path:
    """Recusa gravar dentro da arvore do repositorio.

    Preco de produto e planilha de origem nunca entram no git. A checagem e'
    estrutural (comparacao de caminho resolvido), nao um pedido de disciplina.
    """
    repo = Path(__file__).resolve().parents[2]
    alvo = destino.resolve()
    if alvo == repo or repo in alvo.parents:
        raise ReferenceImportError(
            "destino do snapshot esta DENTRO do repositorio: dado de preco e "
            "planilha de origem nao podem ser versionados. Escolha um caminho "
            "fora da arvore de trabalho."
        )
    return alvo


def write_local_snapshot(snapshot: dict, destino: Path) -> Path:
    """Grava o snapshot sanitizado em JSON. NUNCA sobrescreve."""
    alvo = assert_outside_repo(destino)
    if alvo.exists():
        raise ReferenceImportError(
            f"{alvo.name} ja existe: snapshots sao append-only e nao se "
            f"sobrescrevem. Escolha outro caminho."
        )
    alvo.parent.mkdir(parents=True, exist_ok=True)

    def serial(v):
        if isinstance(v, Decimal):
            return str(v)
        if isinstance(v, datetime):
            return v.isoformat()
        raise TypeError(f"tipo nao serializavel: {type(v).__name__}")

    with alvo.open("w", encoding="utf-8") as fh:
        json.dump(snapshot, fh, ensure_ascii=False, indent=2, default=serial)
    return alvo


# ---------------------------------------------------------------------------
# Publicacao — somente com --apply
# ---------------------------------------------------------------------------

def _get_neon_url() -> str:
    url = os.environ.get("DATABASE_URL")
    if not url:
        raise ReferenceImportError(
            "DATABASE_URL nao definido: o destino do snapshot e' o Neon."
        )
    return url


def publish_snapshot(neon_conn, snapshot: dict) -> dict:
    """UMA transacao: lock -> staging pg_temp -> validacao -> INSERT -> EXCEPT.

    Nao ha DELETE em nenhum ponto: a tabela e' append-only. Se o `snapshot_id` ja
    existir, a publicacao e' RECUSADA — reimportar exige um snapshot novo, nunca
    a sobrescrita de um anterior.
    """
    from psycopg2.extras import execute_values  # noqa: PLC0415 — import tardio

    resultado = {"table": TARGET_TABLE, "inserted": 0, "checks": {}}
    cur = neon_conn.cursor()
    try:
        cur.execute(f"SET LOCAL statement_timeout = '{TARGET_STATEMENT_TIMEOUT}'")
        cur.execute(f"SET LOCAL lock_timeout = '{LOCK_TIMEOUT}'")
        cur.execute("SELECT pg_advisory_xact_lock(%s)", (ADVISORY_LOCK_KEY,))

        cur.execute(
            f"SELECT count(*) AS n FROM {TARGET_TABLE} WHERE snapshot_id = %s",
            (snapshot["snapshot_id"],),
        )
        existentes = cur.fetchone()["n"]
        if existentes:
            raise ReferenceImportError(
                f"snapshot_id ja existe no destino com {existentes} linhas: a "
                f"tabela e' append-only e nada sera sobrescrito."
            )

        cur.execute(f"""
            CREATE TEMP TABLE {STAGING_TABLE}
                (LIKE {TARGET_TABLE} INCLUDING DEFAULTS)
            ON COMMIT DROP
        """)

        cols = list(BUSINESS_COLUMNS) + list(AUDIT_COLUMNS)
        execute_values(
            cur,
            f"INSERT INTO {STAGING_TABLE} ({', '.join(cols)}) VALUES %s",
            [tuple(r[c] for c in cols) for r in snapshot["rows"]],
            page_size=INSERT_PAGE_SIZE,
        )

        cur.execute(f"SELECT count(*) AS n FROM {STAGING_TABLE}")
        na_staging = cur.fetchone()["n"]
        if na_staging != snapshot["total_rows"]:
            raise ReferenceImportError(
                f"staging divergiu da leitura: {na_staging} linhas na staging "
                f"contra {snapshot['total_rows']} lidas."
            )

        cur.execute(f"""
            INSERT INTO {TARGET_TABLE} ({', '.join(cols)})
            SELECT {', '.join(cols)} FROM {STAGING_TABLE}
        """)
        resultado["inserted"] = cur.rowcount

        escopo = f"(SELECT * FROM {TARGET_TABLE} WHERE snapshot_id = %(sid)s)"
        cur.execute(
            f"SELECT count(*) AS n FROM ("
            f"  SELECT {', '.join(cols)} FROM {STAGING_TABLE}"
            f"  EXCEPT SELECT {', '.join(cols)} FROM {escopo} t"
            f") d",
            {"sid": snapshot["snapshot_id"]},
        )
        staging_menos_destino = cur.fetchone()["n"]
        cur.execute(
            f"SELECT count(*) AS n FROM ("
            f"  SELECT {', '.join(cols)} FROM {escopo} t"
            f"  EXCEPT SELECT {', '.join(cols)} FROM {STAGING_TABLE}"
            f") d",
            {"sid": snapshot["snapshot_id"]},
        )
        destino_menos_staging = cur.fetchone()["n"]
        if staging_menos_destino or destino_menos_staging:
            raise ReferenceImportError(
                f"EXCEPT bidirecional divergiu: staging-destino="
                f"{staging_menos_destino} destino-staging={destino_menos_staging}."
            )

        resultado["checks"] = {
            "staging_rows": na_staging,
            "except_both_ways": (staging_menos_destino, destino_menos_staging),
        }
        neon_conn.commit()
        return resultado
    except Exception:
        neon_conn.rollback()
        raise
    finally:
        cur.close()


def _neon_writable(url: str):
    import psycopg2  # noqa: PLC0415 — import tardio: nada conecta no import
    from psycopg2.extras import RealDictCursor  # noqa: PLC0415

    return psycopg2.connect(
        url, cursor_factory=RealDictCursor, connect_timeout=CONNECT_TIMEOUT_SECONDS
    )


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="pma-reference-import",
        description=(
            "Importa um snapshot sanitizado do PRECO SUGERIDO DE REVENDA (PDV) "
            "das tabelas B2B. Le somente a tabela de produtos (linha 33+); o "
            "bloco cadastral das planilhas nunca e' lido nem persistido."
        ),
    )
    p.add_argument("--file", action="append", required=True, metavar="XLSX",
                   help="arquivo .xlsx de origem (repetivel, um por marca)")
    p.add_argument("--out", required=True, metavar="JSON",
                   help="destino do snapshot sanitizado, FORA do repositorio")
    p.add_argument("--run-id", default=None)
    p.add_argument("--apply", action="store_true",
                   help="escreve no Neon. Sem esta flag, nada e' escrito em banco.")
    return p


def _print_report(snapshot: dict, aplicado: dict | None, destino: Path) -> None:
    print("[pma-reference-import]")
    print(f"  snapshot_id     : {snapshot['snapshot_id']}")
    print(f"  captured_at     : {snapshot['captured_at'].isoformat()}")
    print(f"  reference_type  : {snapshot['reference_type']}  (PDV, NAO PMA)")
    print(f"  validity_status : {snapshot['validity_status']}  (origem sem vigencia)")
    print(f"  policy_status   : {snapshot['policy_status']}")
    print(f"  snapshot local  : {destino.name}  (fora do repositorio)")
    print(f"  total de linhas : {snapshot['total_rows']}")
    print("  por marca:")
    for marca in sorted(snapshot["by_brand"]):
        info = snapshot["by_brand"][marca]
        print(f"    {marca:<10} linhas={info['rows']:>4}  com_EAN={info['with_gtin']:>4}"
              f"  com_PDV={info['usable']:>4}  qualidade={info['quality']}")
    print("  exclusoes por arquivo:")
    for rel in snapshot["files"]:
        print(f"    {rel['brand']:<10} aba={rel['sheet']!r} lidas={rel['rows_read']}")
        for motivo, n in sorted(rel["exclusions"].items()):
            print(f"      - {motivo}: {n}")
    print(f"  colunas ignoradas: {', '.join(IGNORED_COLUMNS)}")
    if aplicado is None:
        print("  ESCRITA: nenhuma (sem --apply).")
    else:
        print(f"  ESCRITA: {aplicado['inserted']} linhas em {aplicado['table']}; "
              f"EXCEPT bidirecional={aplicado['checks']['except_both_ways']}")


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    run_id = sanitize_run_id(args.run_id) if args.run_id else default_run_id()
    captured_at = datetime.now(timezone.utc)

    arquivos = [Path(f) for f in args.file]
    for caminho in arquivos:
        if not caminho.is_file():
            print(f"ERRO: arquivo nao encontrado: {caminho.name}", file=sys.stderr)
            return 2

    try:
        destino = assert_outside_repo(Path(args.out))
        snapshot = build_snapshot(arquivos, captured_at, run_id)
        gravado = write_local_snapshot(snapshot, destino)

        aplicado = None
        if args.apply:
            conn = _neon_writable(_get_neon_url())
            try:
                aplicado = publish_snapshot(conn, snapshot)
            finally:
                conn.close()

        _print_report(snapshot, aplicado, gravado)
        return 0
    except (ReferenceImportError, ReferenceContractError) as exc:
        print(f"ERRO: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
