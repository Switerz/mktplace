from datetime import date

import openpyxl
import pytest

from pipelines.ingestion.shopee_raw import inventory as inv

ORDERS_HEADER = [
    "ID do pedido", "Status do pedido", "Data de criação do pedido", "Quantidade",
    "Subtotal do produto", "Total global", "Taxa de comissão líquida",
    "Taxa de serviço líquida", "Valor estimado do frete",
    "Nome de usuário (comprador)", "Cidade", "Cidade",
]


def _write_orders_xlsx(path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "orders"
    ws.append(ORDERS_HEADER)
    for row in rows:
        ws.append(row)
    wb.save(path)


def _write_shop_stats_xlsx(path, daily_rows):
    header = [
        "Data", "Vendas (BRL)", "Vendas Sem os Descontos da Shopee", "Pedidos",
        "Vendas por Pedido", "Cliques Por Produto", "Visitantes",
        "Taxa de Conversão de Pedidos", "Pedidos Cancelados", "Vendas Canceladas",
        "Pedidos Devolvidos / Reembolsados", "Vendas Devolvidas / Reembolsadas",
        "# de compradores", "# de novos compradores", "# de compradores existentes",
        "# de compradores em potencial", "Repetir Índice de Compras",
    ]
    total_row = ["01/01/2026-31/01/2026", "100,00"] + [None] * (len(header) - 2)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    ws.append(total_row)
    ws.append([None] * len(header))
    ws.append(header)
    for row in daily_rows:
        ws.append(row)
    wb.save(path)


def _write_ads_csv(path, ad_rows, encoding="utf-8-sig"):
    lines = [
        "Relatório de Todos os Anúncios CPC - Shopee Brasil",
        "Nome de Usuário,marca",
        "Nome da loja,Marca Loja",
        "ID da Loja,123",
        "Data de Criação do Relatório,19/06/2026 17:12",
        "Período,01/01/2026 - 31/03/2026",
        "",
        "#,Nome do Anúncio,Status,Impressões,Cliques,Despesas,GMV",
    ]
    for i, row in enumerate(ad_rows, start=1):
        lines.append(f"{i}," + ",".join(str(v) for v in row))
    content = "\n".join(lines) + "\n"
    with open(path, "w", encoding=encoding, newline="") as f:
        f.write(content)


# --- detect_source_type / parsing utilitário -------------------------------

def test_detect_source_type_orders():
    assert inv.detect_source_type("Order.all.20260101_20260131.xlsx") == inv.SOURCE_ORDERS


def test_detect_source_type_shop_stats():
    assert inv.detect_source_type("kokeshicosmeticos.shopee-shop-stats.20260101-20260131.xlsx") == inv.SOURCE_SHOP_STATS


def test_detect_source_type_ads():
    assert inv.detect_source_type("Dados+Gerais-01-01-19-03.csv") == inv.SOURCE_ADS


def test_detect_source_type_desconhecido():
    assert inv.detect_source_type("relatorio_qualquer.pdf") == inv.SOURCE_UNKNOWN


def test_parse_filename_part_com_partes():
    group, idx, total = inv.parse_filename_part("Order.all.20260101_20260131_part_2_of_6.xlsx")
    assert (idx, total) == (2, 6)
    assert "_part_" not in group


def test_parse_filename_part_sem_partes():
    group, idx, total = inv.parse_filename_part("Order.all.20260101_20260131.xlsx")
    assert (idx, total) == (None, None)
    assert group == "Order.all.20260101_20260131.xlsx"


def test_parse_filename_date_range_underscore():
    d_from, d_to = inv.parse_filename_date_range("Order.all.20260101_20260131.xlsx")
    assert (d_from, d_to) == (date(2026, 1, 1), date(2026, 1, 31))


def test_parse_filename_date_range_dash():
    d_from, d_to = inv.parse_filename_date_range("brand.shopee-shop-stats.20260201-20260228.xlsx")
    assert (d_from, d_to) == (date(2026, 2, 1), date(2026, 2, 28))


def test_parse_filename_date_range_desconhecido():
    assert inv.parse_filename_date_range("Dados+Gerais-01-01-19-03.csv") == (None, None)


# --- leitura de orders -------------------------------------------------------

def test_read_orders_file_grao_por_linha_de_sku(tmp_path):
    path = tmp_path / "Order.all.20260101_20260131.xlsx"
    _write_orders_xlsx(
        path,
        [
            ["1001", "Concluído", "2026-01-05", 2, 50.0, 120.0, 5.0, 2.0, 10.0, "user_a", "SP", "SP"],
            ["1001", "Concluído", "2026-01-05", 1, 30.0, 120.0, 5.0, 2.0, 10.0, "user_a", "SP", "SP"],
            ["1002", "Cancelado", "2026-01-06", 1, 20.0, 20.0, 0.0, 0.0, 0.0, "user_b", "RJ", "RJ"],
        ],
    )
    result = inv.read_orders_file(path)
    assert len(result.rows) == 3  # duas linhas de SKU do pedido 1001 + uma do 1002 — nada é agregado
    assert result.headers == ORDERS_HEADER
    assert result.header_row_index == 0


def test_read_orders_file_linha_vazia_e_rejeitada_nao_perdida(tmp_path):
    path = tmp_path / "Order.all.x.xlsx"
    _write_orders_xlsx(
        path,
        [
            ["1001", "Concluído", "2026-01-05", 2, 50.0, 120.0, 5.0, 2.0, 10.0, "user_a", "SP", "SP"],
            [None] * len(ORDERS_HEADER),
            ["1002", "Cancelado", "2026-01-06", 1, 20.0, 20.0, 0.0, 0.0, 0.0, "user_b", "RJ", "RJ"],
        ],
    )
    result = inv.read_orders_file(path)
    assert len(result.rows) == 2
    assert len(result.rejects) == 1
    # reconciliação exata: físicas == parseadas + rejeitadas
    assert len(result.rows) + len(result.rejects) == 3


def test_read_orders_file_header_duplicado_nao_sobrescreve(tmp_path):
    path = tmp_path / "Order.all.dup.xlsx"
    _write_orders_xlsx(
        path,
        [["1001", "Concluído", "2026-01-05", 2, 50.0, 120.0, 5.0, 2.0, 10.0, "user_a", "SP-1", "SP-2"]],
    )
    result = inv.read_orders_file(path)
    payload = result.rows[0].raw_payload
    # "Cidade" aparece 2x no header — a segunda ocorrência não pode sobrescrever a primeira
    assert payload["Cidade"] == "SP-1"
    assert payload["Cidade__col11"] == "SP-2"


def test_read_orders_file_cancelado_nao_e_descartado(tmp_path):
    """Raw não filtra por status — pedidos cancelados continuam presentes."""
    path = tmp_path / "Order.all.cancel.xlsx"
    _write_orders_xlsx(
        path,
        [["1002", "Cancelado", "2026-01-06", 1, 20.0, 20.0, 0.0, 0.0, 0.0, "user_b", "RJ", "RJ"]],
    )
    result = inv.read_orders_file(path)
    assert len(result.rows) == 1
    assert result.rows[0].raw_payload["Status do pedido"] == "Cancelado"


def test_read_orders_file_vazio_sem_linhas(tmp_path):
    path = tmp_path / "Order.all.empty.xlsx"
    wb = openpyxl.Workbook()
    wb.save(path)
    result = inv.read_orders_file(path)
    assert result.rows == []
    assert result.headers == []


def test_read_orders_file_corrompido_levanta_source_read_error(tmp_path):
    path = tmp_path / "Order.all.corrupto.xlsx"
    path.write_bytes(b"isso nao e um xlsx valido")
    with pytest.raises(inv.SourceReadError):
        inv.read_orders_file(path)


# --- leitura de shop-stats ----------------------------------------------------

def test_read_shop_stats_inclui_linha_de_total_e_linhas_diarias(tmp_path):
    path = tmp_path / "brand.shopee-shop-stats.20260101-20260131.xlsx"
    daily = [
        ["01/01/2026"] + [1] * 16,
        ["02/01/2026"] + [2] * 16,
    ]
    _write_shop_stats_xlsx(path, daily)
    result = inv.read_shop_stats_file(path)
    # 1 linha de total do período + 2 linhas diárias
    assert len(result.rows) == 3
    assert result.rows[0].source_row_number == 1  # linha de total
    assert result.rows[1].source_row_number == 4  # primeira linha diária


def test_read_shop_stats_template_muito_curto_levanta_erro(tmp_path):
    path = tmp_path / "brand.shopee-shop-stats.curto.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Data", "Vendas (BRL)"])
    wb.save(path)
    with pytest.raises(inv.SourceReadError):
        inv.read_shop_stats_file(path)


# --- leitura de ads -----------------------------------------------------------

def test_read_ads_file_grao_por_anuncio(tmp_path):
    path = tmp_path / "Dados+Gerais.csv"
    _write_ads_csv(path, [["Anúncio 1", "Ativo", 1000, 50, "10,00", "100,00"], ["Anúncio 2", "Ativo", 2000, 80, "20,00", "200,00"]])
    result = inv.read_ads_file(path)
    assert len(result.rows) == 2
    assert result.encoding == "utf-8-sig"
    assert result.rows[0].raw_payload["Nome do Anúncio"] == "Anúncio 1"


def test_read_ads_file_encoding_fallback_latin1(tmp_path):
    path = tmp_path / "Dados+Gerais.csv"
    _write_ads_csv(path, [["Anúncio Ção", "Ativo", 1000, 50, "10,00", "100,00"]], encoding="latin-1")
    result = inv.read_ads_file(path)
    assert result.encoding in ("cp1252", "latin-1")
    assert len(result.rows) == 1


def test_read_ads_file_sem_header_levanta_erro(tmp_path):
    path = tmp_path / "Dados+SemHeader.csv"
    path.write_text("linha,sem,cabecalho,valido\n1,2,3,4\n", encoding="utf-8-sig")
    with pytest.raises(inv.SourceReadError):
        inv.read_ads_file(path)


# --- scan_file / scan_directory -----------------------------------------------

def _make_shopee_tree(tmp_path):
    base = tmp_path / "shopee"
    (base / "apice").mkdir(parents=True)
    (base / "barbours").mkdir(parents=True)
    (base / "unknown_brand").mkdir(parents=True)
    return base


def test_scan_file_marca_brand_desconhecida(tmp_path):
    base = _make_shopee_tree(tmp_path)
    path = base / "unknown_brand" / "Order.all.20260101_20260131.xlsx"
    _write_orders_xlsx(path, [["1", "Concluído", "2026-01-01", 1, 10.0, 10.0, 0, 0, 0, "u", "SP", "SP"]])
    record = inv.scan_file(path, base)
    assert record.brand == "unknown_brand"
    assert record.brand_known is False


def test_scan_file_caminho_relativo_nunca_absoluto(tmp_path):
    base = _make_shopee_tree(tmp_path)
    path = base / "apice" / "Order.all.20260101_20260131.xlsx"
    _write_orders_xlsx(path, [["1", "Concluído", "2026-01-01", 1, 10.0, 10.0, 0, 0, 0, "u", "SP", "SP"]])
    record = inv.scan_file(path, base)
    assert record.relative_path == "apice/Order.all.20260101_20260131.xlsx"
    assert not record.relative_path.startswith("/")
    assert ":" not in record.relative_path
    assert str(tmp_path) not in record.relative_path


def test_scan_file_vazio(tmp_path):
    base = _make_shopee_tree(tmp_path)
    path = base / "apice" / "Order.all.vazio.xlsx"
    path.touch()
    record = inv.scan_file(path, base)
    assert record.is_empty is True
    assert record.is_readable is False


def test_scan_file_corrompido_nao_derruba_inventario(tmp_path):
    base = _make_shopee_tree(tmp_path)
    path = base / "apice" / "Order.all.corrupto.xlsx"
    path.write_bytes(b"lixo binario")
    record = inv.scan_file(path, base)
    assert record.is_readable is False
    assert record.error_message


def test_scan_file_desconhecido_nao_tenta_parsear_conteudo(tmp_path):
    base = _make_shopee_tree(tmp_path)
    path = base / "apice" / "relatorio.pdf"
    path.write_bytes(b"%PDF-1.4 fake")
    record = inv.scan_file(path, base)
    assert record.source_type == inv.SOURCE_UNKNOWN
    assert record.headers is None
    assert record.file_sha256 is not None  # hash de arquivo continua sendo feito


def test_scan_directory_encontra_todos_os_tipos(tmp_path):
    base = _make_shopee_tree(tmp_path)
    _write_orders_xlsx(
        base / "apice" / "Order.all.20260101_20260131.xlsx",
        [["1", "Concluído", "2026-01-01", 1, 10.0, 10.0, 0, 0, 0, "u", "SP", "SP"]],
    )
    _write_shop_stats_xlsx(base / "apice" / "apice.shopee-shop-stats.20260101-20260131.xlsx", [["01/01/2026"] + [1] * 16])
    _write_ads_csv(base / "apice" / "Dados+Gerais.csv", [["Anúncio 1", "Ativo", 1000, 50, "10,00", "100,00"]])

    records = inv.scan_directory(base)
    source_types = {r.source_type for r in records}
    assert source_types == {inv.SOURCE_ORDERS, inv.SOURCE_SHOP_STATS, inv.SOURCE_ADS}


def test_scan_directory_diretorio_inexistente_levanta_erro(tmp_path):
    with pytest.raises(FileNotFoundError):
        inv.scan_directory(tmp_path / "nao_existe")


# --- idempotência / duplicados / schema drift / overlap -----------------------

def test_scan_file_e_idempotente_mesmo_arquivo(tmp_path):
    base = _make_shopee_tree(tmp_path)
    path = base / "apice" / "Order.all.20260101_20260131.xlsx"
    _write_orders_xlsx(path, [["1", "Concluído", "2026-01-01", 1, 10.0, 10.0, 0, 0, 0, "u", "SP", "SP"]])
    r1 = inv.scan_file(path, base)
    r2 = inv.scan_file(path, base)
    assert r1.file_sha256 == r2.file_sha256
    assert r1.schema_fingerprint == r2.schema_fingerprint


def test_find_duplicate_files_por_sha256(tmp_path):
    import shutil

    base = _make_shopee_tree(tmp_path)
    rows = [["1", "Concluído", "2026-01-01", 1, 10.0, 10.0, 0, 0, 0, "u", "SP", "SP"]]
    p1 = base / "apice" / "Order.all.a.xlsx"
    p2 = base / "barbours" / "Order.all.b.xlsx"
    _write_orders_xlsx(p1, rows)
    # cópia byte-a-byte: gerar dois xlsx "iguais" via openpyxl pode produzir
    # bytes diferentes (metadados internos com timestamp de criação)
    shutil.copyfile(p1, p2)
    records = [inv.scan_file(p1, base), inv.scan_file(p2, base)]
    dups = inv.find_duplicate_files(records)
    assert len(dups) == 1
    assert set(next(iter(dups.values()))) == {"apice/Order.all.a.xlsx", "barbours/Order.all.b.xlsx"}


def test_find_schema_drift_entre_marcas(tmp_path):
    base = _make_shopee_tree(tmp_path)
    p1 = base / "apice" / "Order.all.a.xlsx"
    p2 = base / "barbours" / "Order.all.b.xlsx"
    _write_orders_xlsx(p1, [["1", "Concluído", "2026-01-01", 1, 10.0, 10.0, 0, 0, 0, "u", "SP", "SP"]])
    # segundo arquivo com um header a mais -> fingerprint diferente
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(ORDERS_HEADER + ["Coluna Extra"])
    ws.append(["1", "Concluído", "2026-01-01", 1, 10.0, 10.0, 0, 0, 0, "u", "SP", "SP", "x"])
    wb.save(p2)

    records = [inv.scan_file(p1, base), inv.scan_file(p2, base)]
    drift = inv.find_schema_drift(records)
    assert inv.SOURCE_ORDERS in drift
    assert len(drift[inv.SOURCE_ORDERS]) == 2


def test_find_overlapping_exports_detecta_sobreposicao(tmp_path):
    base = _make_shopee_tree(tmp_path)
    rows = [["1", "Concluído", "2026-01-01", 1, 10.0, 10.0, 0, 0, 0, "u", "SP", "SP"]]
    p1 = base / "apice" / "Order.all.20260101_20260131.xlsx"
    p2 = base / "apice" / "Order.all.20260115_20260215.xlsx"
    _write_orders_xlsx(p1, rows)
    _write_orders_xlsx(p2, rows)
    records = [inv.scan_file(p1, base), inv.scan_file(p2, base)]
    overlaps = inv.find_overlapping_exports(records)
    assert len(overlaps) == 1


def test_find_overlapping_exports_nao_conta_partes_do_mesmo_export(tmp_path):
    base = _make_shopee_tree(tmp_path)
    rows = [["1", "Concluído", "2026-01-01", 1, 10.0, 10.0, 0, 0, 0, "u", "SP", "SP"]]
    p1 = base / "apice" / "Order.all.20260101_20260131_part_1_of_2.xlsx"
    p2 = base / "apice" / "Order.all.20260101_20260131_part_2_of_2.xlsx"
    _write_orders_xlsx(p1, rows)
    _write_orders_xlsx(p2, rows)
    records = [inv.scan_file(p1, base), inv.scan_file(p2, base)]
    overlaps = inv.find_overlapping_exports(records)
    assert overlaps == []


def test_build_summary_agrega_tudo(tmp_path):
    base = _make_shopee_tree(tmp_path)
    _write_orders_xlsx(
        base / "apice" / "Order.all.20260101_20260131.xlsx",
        [["1", "Concluído", "2026-01-01", 1, 10.0, 10.0, 0, 0, 0, "u", "SP", "SP"]],
    )
    records = inv.scan_directory(base)
    summary = inv.build_summary(records)
    assert summary["total_files"] == len(records)
    assert "orders" in summary["by_source_type"]
